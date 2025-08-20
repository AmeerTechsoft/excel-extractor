#!/usr/bin/env python3
"""
OpenPyXL-based extractor for cross-platform environments without local Excel.
Provides a similar API to ExcelFormulaExtractor using openpyxl.
Limitations:
- No live calculation engine; formulas are returned as text, display_text may be None
- Data validation list resolution is best-effort for ranges, named ranges, and simple literals
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter


class OpenpyxlExcelExtractor:
	"""Extract Excel metadata using openpyxl (cross-platform)."""

	def __init__(self, excel_file_path: str):
		self.excel_file_path = Path(excel_file_path)
		self.workbook = None
		self.worksheet: Optional[Worksheet] = None
		if not self.excel_file_path.exists():
			raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

	def __enter__(self):
		self.open_workbook()
		return self

	def __exit__(self, exc_type, exc_val, exc_tb):
		self.close_workbook()

	def open_workbook(self) -> None:
		# data_only=False ensures formulas remain in cell.value for extraction
		self.workbook = load_workbook(filename=str(self.excel_file_path), data_only=False, read_only=False)

	def close_workbook(self) -> None:
		try:
			if self.workbook:
				self.workbook.close()
		except Exception:
			pass

	def _select_sheet(self, sheet_name: Optional[str]) -> Worksheet:
		if sheet_name:
			self.worksheet = self.workbook[sheet_name]
		else:
			self.worksheet = self.workbook.active
		return self.worksheet

	def get_worksheet_info(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
		ws = self._select_sheet(sheet_name)
		# calculate_dimension returns a range like A1:D10 (best effort)
		try:
			dimension = ws.calculate_dimension() or "A1:A1"
		except Exception:
			dimension = "A1:A1"
		try:
			min_row, min_col, max_row, max_col = ws.calculate_dimension().split(":")[0], None, None, None
		except Exception:
			pass
		min_row_val, min_col_val, max_row_val, max_col_val = ws.min_row, ws.min_column, ws.max_row, ws.max_column
		rows = max(0, (max_row_val or 0))
		cols = max(0, (max_col_val or 0))
		return {
			"sheet_name": ws.title,
			"used_range": dimension,
			"rows": rows,
			"columns": cols,
			"total_cells": rows * cols,
		}

	def _cell_basic_format(self, cell) -> Dict[str, Any]:
		fmt: Dict[str, Any] = {}
		try:
			fmt["number_format"] = cell.number_format
		except Exception:
			fmt["number_format"] = None
		try:
			fmt["font_name"] = getattr(cell.font, "name", None)
			fmt["font_size"] = getattr(cell.font, "size", None)
			fmt["font_bold"] = getattr(cell.font, "bold", None)
			fmt["font_italic"] = getattr(cell.font, "italic", None)
		except Exception:
			pass
		try:
			fill = cell.fill
			color = getattr(fill, "fgColor", None)
			if color is not None and getattr(color, "rgb", None):
				rgb = color.rgb
				fmt["fill_color"] = {"rgb": {"r": int(rgb[0:2], 16), "g": int(rgb[2:4], 16), "b": int(rgb[4:6], 16)}}
		except Exception:
			pass
		try:
			align = cell.alignment
			fmt["horizontal_alignment"] = getattr(align, "horizontal", None)
			fmt["vertical_alignment"] = getattr(align, "vertical", None)
		except Exception:
			pass
		# Merged cells info
		try:
			merged = False
			merge_area = None
			for cr in self.worksheet.merged_cells.ranges:
				if cell.coordinate in cr:
					merged = True
					merge_area = str(cr)
					break
			fmt["merged"] = merged
			fmt["merge_area"] = merge_area
		except Exception:
			pass
		return fmt

	def _cell_hyperlink(self, cell) -> Optional[Dict[str, Any]]:
		try:
			hl = cell.hyperlink
			if hl is None:
				return None
			return {
				"address": hl.target,
				"text_to_display": getattr(hl, "display", None),
			}
		except Exception:
			return None

	def _cell_note(self, cell) -> Optional[str]:
		try:
			c = cell.comment
			return c.text if c else None
		except Exception:
			return None

	def _data_validation_for_cell(self, ws: Worksheet, row: int, col: int) -> Optional[Dict[str, Any]]:
		try:
			if not ws.data_validations:
				return None
			coord = f"{get_column_letter(col)}{row}"
			for dv in ws.data_validations.dataValidation:  # type: DataValidation
				try:
					for sq in dv.sqref:
						cr = CellRange(sq)
						if coord in cr:
							v: Dict[str, Any] = {
								"type": dv.type,
								"type_name": dv.type,
								"operator": dv.operator,
								"ignore_blank": dv.allow_blank,
								"in_cell_dropdown": True,  # openpyxl implies list validations use dropdown
								"formula1": dv.formula1,
								"formula2": dv.formula2,
							}
							# Resolve list items if possible
							if str(dv.type).lower() == "list":
								resolved = self._resolve_validation_list_items(ws, dv.formula1)
								if resolved is not None:
									v["list_items"] = resolved
							return v
				except Exception:
					continue
			return None
		except Exception:
			return None

	def _flatten_values(self, vals: Any) -> List[Any]:
		out: List[Any] = []
		if vals is None:
			return out
		if isinstance(vals, list):
			for r in vals:
				if isinstance(r, list):
					out.extend([v for v in r if v is not None])
				else:
					out.append(r)
		else:
			out.append(vals)
		return out

	def _values_from_range(self, ws: Worksheet, ref: str) -> Optional[List[Any]]:
		try:
			cr = CellRange(ref)
			vals: List[Any] = []
			for r in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
				vals.extend([c.value for c in r])
			return self._flatten_values(vals)
		except Exception:
			return None

	def _resolve_named_range(self, wb, name: str) -> Optional[List[Any]]:
		try:
			dn = wb.defined_names.get(name)
			if dn is None:
				return None
			# A defined name can have destinations across sheets
			for title, coord in dn.destinations:
				ws = wb[title]
				vals = self._values_from_range(ws, coord)
				if vals is not None:
					return vals
			return None
		except Exception:
			return None

	def _resolve_table_column(self, wb, ref: str) -> Optional[List[Any]]:
		# Basic parser for TableName[Column]
		try:
			if "[" not in ref or "]" not in ref:
				return None
			table_name = ref.split("[", 1)[0]
			col_name = ref.split("[", 1)[1].split("]", 1)[0]
			for ws in wb.worksheets:
				try:
					if table_name in getattr(ws, "tables", {}):
						tbl = ws.tables[table_name]
						# Table ref like A1:D20; need to find column index by header
						ref_range = CellRange(tbl.ref)
						# Find header row (top row of table)
						headers = [c.value for c in next(ws.iter_rows(min_row=ref_range.min_row, max_row=ref_range.min_row, min_col=ref_range.min_col, max_col=ref_range.max_col))]
						if col_name in headers:
							col_idx = headers.index(col_name) + ref_range.min_col
							data_min_row = ref_range.min_row + 1
							vals = [c.value for c in ws.iter_rows(min_row=data_min_row, max_row=ref_range.max_row, min_col=col_idx, max_col=col_idx, values_only=True)]
							# iter_rows with values_only returns tuples; handle it
							flat: List[Any] = []
							for t in vals:
								if isinstance(t, tuple) and t:
									flat.append(t[0])
								else:
									flat.append(t)
							return flat
				except Exception:
					continue
			return None
		except Exception:
			return None

	def _resolve_validation_list_items(self, ws: Worksheet, formula1: Optional[str]) -> Optional[List[Any]]:
		if not formula1:
			return None
		try:
			f = str(formula1).strip()
			if f.startswith("="):
				f = f[1:]
			if f.startswith('"') and f.endswith('"') and len(f) >= 2:
				f = f[1:-1]
			# Literal CSV list
			if "," in f and "!" not in f and ":" not in f and "[" not in f:
				return [s.strip() for s in f.split(",")]
			# Structured table reference
			if "[" in f and "]" in f:
				resolved = self._resolve_table_column(ws.parent, f)
				if resolved is not None:
					return resolved
			# Sheet range like Sheet1!$A$1:$A$10
			if "!" in f or ":" in f:
				if "!" in f:
					sheet_name, addr = f.split("!", 1)
					ws_target = ws.parent[sheet_name.strip().strip("'")]
				else:
					ws_target = ws
					addr = f
				vals = self._values_from_range(ws_target, addr)
				if vals is not None:
					return vals
			# Named range
			resolved = self._resolve_named_range(ws.parent, f)
			if resolved is not None:
				return resolved
		except Exception:
			return None
		return None

	def _column_to_letter(self, col: int) -> str:
		return get_column_letter(col)

	def _extract_cell_full_details(self, cell) -> Dict[str, Any]:
		value = cell.value
		formula = value if isinstance(value, str) and value.startswith("=") else None
		return {
			"address": cell.coordinate,
			"row": cell.row,
			"column": cell.column,
			"column_letter": self._column_to_letter(cell.column),
			"value": None if formula else value,
			"formula": formula,
			"display_text": None,  # openpyxl does not render display text
			"format": self._cell_basic_format(cell),
			"hyperlink": self._cell_hyperlink(cell),
			"note": self._cell_note(cell),
			"data_validation": self._data_validation_for_cell(self.worksheet, cell.row, cell.column),
		}

	def extract_sheet_full_details(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
		ws = self._select_sheet(sheet_name)
		info = self.get_worksheet_info(sheet_name)
		cells: List[Dict[str, Any]] = []
		if ws.max_row and ws.max_column:
			for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
				for cell in row:
					try:
						cells.append(self._extract_cell_full_details(cell))
					except Exception:
						continue
		# Tables
		tables: List[Dict[str, Any]] = []
		try:
			for name, tbl in getattr(ws, "tables", {}).items():
				try:
					# openpyxl Table has displayName property
					tables.append({
						"name": getattr(tbl, "displayName", name),
						"range": getattr(tbl, "ref", None),
						"data_body_range": None,
						"header_row_range": None,
						"totals_row_range": None,
						"show_totals": getattr(tbl, "showTotalsRow", False),
						"columns": [c.name for c in getattr(tbl, "tableColumns", [])] if hasattr(tbl, "tableColumns") else [],
					})
			except Exception:
				pass
		return {
			"sheet": {
				"name": info.get("sheet_name"),
				"used_range": info.get("used_range"),
				"rows": info.get("rows"),
				"columns": info.get("columns"),
				"visible": True,
				"protect_contents": None,
				"protect_drawing": None,
				"protect_scenarios": None,
			},
			"cells": cells,
			"tables": tables,
		}

	def extract_workbook_full_details(self) -> Dict[str, Any]:
		all_sheets: List[Dict[str, Any]] = []
		for ws in self.workbook.worksheets:
			self.worksheet = ws
			try:
				all_sheets.append(self.extract_sheet_full_details(ws.title))
			except Exception as e:
				all_sheets.append({"sheet": {"name": ws.title, "error": str(e)}, "cells": [], "tables": []})
		# Named ranges metadata
		names: List[Dict[str, Any]] = []
		try:
			for name in self.workbook.defined_names.definedName:
				names.append({"name": name.name, "refers_to": name.attr_text})
		except Exception:
			pass
		return {
			"file_path": str(self.excel_file_path),
			"extraction_timestamp": datetime.now().isoformat(),
			"workbook": {
				"sheet_count": len(self.workbook.worksheets),
				"sheets": all_sheets,
				"names": names,
			},
		}

	def extract_formulas_from_range(self, start_cell: str = "A1", end_cell: Optional[str] = None) -> List[Dict[str, Any]]:
		ws = self.worksheet or self._select_sheet(None)
		if end_cell:
			rng = CellRange(f"{start_cell}:{end_cell}")
			cells_iter = ws.iter_rows(min_row=rng.min_row, max_row=rng.max_row, min_col=rng.min_col, max_col=rng.max_col)
		else:
			rng = CellRange(start_cell)
			cells_iter = ws.iter_rows(min_row=rng.min_row, max_row=rng.max_row, min_col=rng.min_col, max_col=rng.max_col)
		out: List[Dict[str, Any]] = []
		for row in cells_iter:
			for cell in row:
				info = self._extract_cell_full_details(cell)
				if info.get("formula") or (isinstance(info.get("value"), (int, float)) and info.get("value") != 0):
					out.append(info)
		return out

	def extract_all_formulas(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
		info = self.get_worksheet_info(sheet_name)
		ws = self.worksheet
		all_formulas: List[Dict[str, Any]] = []
		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
			for cell in row:
				ci = self._extract_cell_full_details(cell)
				if ci.get("formula") or (isinstance(ci.get("value"), (int, float)) and ci.get("value") != 0):
					all_formulas.append(ci)
		return {
			"file_path": str(self.excel_file_path),
			"extraction_timestamp": datetime.now().isoformat(),
			"worksheet_info": info,
			"total_formulas_found": len(all_formulas),
			"formulas": all_formulas,
		}

	def extract_formula_dependencies(self, cell_address: str) -> Dict[str, Any]:
		ws = self.worksheet or self._select_sheet(None)
		cell = ws[cell_address]
		formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
		if not formula:
			return {"error": "Cell does not contain a formula"}
		deps = self._parse_formula_dependencies(formula)
		dependent_values: Dict[str, Any] = {}
		for dep in deps:
			try:
				c = ws[dep]
				dependent_values[dep] = {
					"value": c.value,
					"formula": c.value if isinstance(c.value, str) and c.value.startswith("=") else None,
					"address": dep,
				}
			except Exception:
				dependent_values[dep] = {"error": "Could not access cell"}
		return {
			"cell_address": cell_address,
			"formula": formula,
			"dependencies": deps,
			"dependent_values": dependent_values,
			"calculated_value": None,
		}

	def _parse_formula_dependencies(self, formula: str) -> List[str]:
		import re
		cell_pattern = r"[A-Z]+\d+"
		matches = re.findall(cell_pattern, formula.upper())
		return sorted(list(set(matches)))

	def export_to_json(self, data: Dict[str, Any], output_file: str) -> bool:
		try:
			with open(output_file, 'w', encoding='utf-8') as f:
				json.dump(data, f, indent=2, ensure_ascii=False, default=str)
			return True
		except Exception:
			return False

	def export_to_text(self, data: Dict[str, Any], output_file: str) -> bool:
		try:
			with open(output_file, 'w', encoding='utf-8') as f:
				f.write("EXCEL FORMULA EXTRACTION REPORT\n")
				f.write("=" * 50 + "\n\n")
				f.write(f"File: {data.get('file_path', 'Unknown')}\n")
				f.write(f"Extracted: {data.get('extraction_timestamp', 'Unknown')}\n")
				wsn = (data.get('worksheet_info', {}) or {}).get('sheet_name')
				if wsn:
					f.write(f"Worksheet: {wsn}\n")
				f.write(f"Total Formulas: {data.get('total_formulas_found', 0)}\n\n")
				f.write("FORMULAS:\n")
				f.write("-" * 20 + "\n")
				for formula_info in data.get('formulas', []):
					f.write(f"Cell: {formula_info.get('address', 'Unknown')}\n")
					if formula_info.get('formula'):
						f.write(f"Formula: {formula_info['formula']}\n")
					f.write(f"Value: {formula_info.get('value', 'N/A')}\n")
					f.write(f"Row: {formula_info.get('row', 'N/A')}, Column: {formula_info.get('column_letter', 'N/A')}\n")
					f.write("-" * 10 + "\n")
			return True
		except Exception:
			return False 